#!/usr/bin/env python3
"""
Convert the official Ciqual 2020 Excel file to a Supabase SQL migration.

Usage:
  1. Download the Ciqual 2020 Excel file from:
     https://ciqual.anses.fr/  →  "Télécharger la table Ciqual 2020"
     File name: "Table Ciqual 2020_FR_2020 07 07.xls" (or .xlsx)

  2. Run:
     pip3 install openpyxl xlrd
     python3 ciqual2sql.py "Table Ciqual 2020_FR_2020 07 07.xls"

  3. Execute the output SQL in Supabase SQL Editor:
     psql or Supabase Dashboard → SQL Editor → paste content of 004_ciqual_complete.sql
"""

import sys
import re
import os

def parse_value(val) -> float | None:
    """Parse a Ciqual nutrient value. Returns None for missing/trace values."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('-', '', 'traces', 'Traces', 'nd', 'ND', '-'):
        return None
    # Remove "< " prefix (below detection limit — treat as 0)
    s = re.sub(r'^[<>]\s*', '', s)
    # Replace comma decimal separator
    s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None

def normalize_name(name: str) -> str:
    """Normalize ligatures and apostrophes for consistent search."""
    name = name.replace('Œ', 'Oe').replace('œ', 'oe')
    name = name.replace('Æ', 'Ae').replace('æ', 'ae')
    name = name.replace('’', "'").replace('‘', "'")  # curly apostrophes
    return name.strip()

def escape_sql(s: str) -> str:
    return s.replace("'", "''")

def load_xls(path: str):
    """Load old .xls format using xlrd."""
    import xlrd
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    rows = []
    for r in range(1, ws.nrows):
        row = {}
        for c, h in enumerate(headers):
            row[h] = ws.cell_value(r, c)
        rows.append(row)
    return headers, rows

def load_xlsx(path: str):
    """Load .xlsx format using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else '' for h in next(rows_iter)]
    rows = []
    for row_vals in rows_iter:
        row = {headers[i]: row_vals[i] for i in range(len(headers))}
        rows.append(row)
    return headers, rows

def find_column(headers: list[str], patterns: list[str]) -> str | None:
    """Find the first header matching any of the given patterns (case-insensitive)."""
    for pattern in patterns:
        pat_lower = pattern.lower()
        for h in headers:
            if pat_lower in h.lower():
                return h
    return None

def convert(input_path: str, output_path: str):
    ext = os.path.splitext(input_path)[1].lower()
    print(f"Loading {input_path}...")

    if ext == '.xls':
        headers, rows = load_xls(input_path)
    else:
        headers, rows = load_xlsx(input_path)

    print(f"Columns found: {len(headers)}")
    print(f"Rows found: {len(rows)}")

    # Map column names (Ciqual 2020 uses verbose French names)
    col_code = find_column(headers, ['alim_code', 'code aliment', 'code'])
    col_name = find_column(headers, ['alim_nom_fr', 'nom de l\'aliment', 'alim_nom'])
    col_kcal = find_column(headers, [
        'Energie, Règlement UE N° 1169/2011 (kcal/100 g)',
        'Energie (kcal/100 g)',
        'kcal',
        'énergie',
        'energie',
    ])
    col_prot = find_column(headers, [
        'Protéines, N x facteur de Jones (g/100 g)',
        'Protéines (g/100 g)',
        'protéines',
        'proteines',
        'prot',
    ])
    col_carb = find_column(headers, [
        'Glucides (g/100 g)',
        'glucides',
        'carb',
    ])
    col_fat = find_column(headers, [
        'Lipides (g/100 g)',
        'lipides',
        'fat',
        'matières grasses',
    ])

    print(f"\nColumn mapping:")
    print(f"  code  → {col_code!r}")
    print(f"  name  → {col_name!r}")
    print(f"  kcal  → {col_kcal!r}")
    print(f"  prot  → {col_prot!r}")
    print(f"  carbs → {col_carb!r}")
    print(f"  fat   → {col_fat!r}")

    missing = [k for k, v in {'code': col_code, 'name': col_name, 'kcal': col_kcal,
                               'prot': col_prot, 'carbs': col_carb, 'fat': col_fat}.items() if v is None]
    if missing:
        print(f"\nERROR: Could not find columns for: {missing}")
        print("Available headers:")
        for h in headers:
            print(f"  {h!r}")
        sys.exit(1)

    foods = []
    skipped = 0
    for row in rows:
        code_raw = row.get(col_code)
        name_raw = row.get(col_name)
        kcal_raw = row.get(col_kcal)

        if not code_raw or not name_raw:
            skipped += 1
            continue

        name = normalize_name(str(name_raw))
        kcal = parse_value(kcal_raw)

        if not name or kcal is None:
            skipped += 1
            continue

        try:
            code = int(float(str(code_raw)))
        except (ValueError, TypeError):
            skipped += 1
            continue

        prot = parse_value(row.get(col_prot)) or 0.0
        carb = parse_value(row.get(col_carb)) or 0.0
        fat  = parse_value(row.get(col_fat))  or 0.0

        foods.append((code, name, round(kcal, 1), round(prot, 2), round(carb, 2), round(fat, 2)))

    print(f"\n{len(foods)} foods to import ({skipped} skipped — no name or no calories)")

    # Write SQL
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("-- Ciqual 2020 ANSES — full database import\n")
        f.write("-- Generated by scripts/ciqual2sql.py\n")
        f.write(f"-- {len(foods)} foods\n\n")

        f.write("-- Drop existing data and re-import cleanly\n")
        f.write("TRUNCATE public.reference_foods RESTART IDENTITY CASCADE;\n\n")

        f.write("INSERT INTO public.reference_foods (id, name, kcal, protein_g, carbs_g, fat_g) VALUES\n")
        values = []
        for code, name, kcal, prot, carb, fat in foods:
            values.append(f"({code}, '{escape_sql(name)}', {kcal}, {prot}, {carb}, {fat})")

        f.write(',\n'.join(values))
        f.write(";\n\n")

        f.write("-- Rebuild FTS index\n")
        f.write("REINDEX INDEX reference_foods_name_fts_idx;\n")

    print(f"\nSQL written to: {output_path}")
    print(f"Execute it in Supabase SQL Editor to import all {len(foods)} Ciqual foods.")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'supabase/migrations/004_ciqual_complete.sql'

    if not os.path.exists(input_file):
        print(f"ERROR: File not found: {input_file}")
        sys.exit(1)

    convert(input_file, output_file)
